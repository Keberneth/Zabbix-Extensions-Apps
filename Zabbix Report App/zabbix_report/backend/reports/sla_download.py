"""
backend/reports/sla_download.py

Helper for generating an SLA/SLI Excel workbook entirely in memory.

Usage example:

    from .sla import get_sla_sli
    from .sla_download import create_sla_xlsx_bytes

    data = get_sla_sli(periods=12)
    xlsx_bytes = create_sla_xlsx_bytes(data)
"""

from io import BytesIO
from typing import Any, Dict, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from .sla import get_sla_sli


def _style_header_cell(cell):
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal="center")
    cell.fill = PatternFill(start_color="FFE5E5E5", end_color="FFE5E5E5", fill_type="solid")
    thin = Side(style="thin", color="FFBBBBBB")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _style_body_cell(cell):
    cell.alignment = Alignment(horizontal="center")
    thin = Side(style="thin", color="FFDDDDDD")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def _auto_width(ws):
    """Best-effort auto column width based on cell contents."""
    max_len = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            col = cell.column_letter
            v = str(cell.value)
            max_len[col] = max(max_len.get(col, 0), len(v))
    for col, length in max_len.items():
        ws.column_dimensions[col].width = min(max(length + 2, 10), 40)


def _build_sla_sheet(wb: Workbook, sla_id: str, sla_info: Dict[str, Any]) -> None:
    """
    Create one sheet per SLA.

    Expected sla_info structure (as returned by get_sla_sli):

        {
          "sla_name": str,
          "period_from": int,
          "periods": int,
          "service_data": {
            <serviceid>: {
              "name": str,
              "slo": "99%",
              "month_labels": ["2025-01", ...],
              "monthly_sli": ["100%", "99%", ...]
            },
            ...
          }
        }
    """
    sla_name = sla_info.get("sla_name") or f"SLA {sla_id}"
    title = sla_name[:31] or f"SLA_{sla_id}"  # Excel sheet name limit
    ws = wb.create_sheet(title=title)

    service_data: Dict[Any, Any] = sla_info.get("service_data") or {}
    # Collect month labels from first service that has them
    month_labels = []
    for svc in service_data.values():
        labels = svc.get("month_labels") or []
        if labels:
            month_labels = labels
            break

    # Header
    headers = ["Service name", "SLO"] + month_labels
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        _style_header_cell(cell)

    # Rows per service
    row_idx = 2
    for svc in service_data.values():
        name = svc.get("name", "")
        slo = svc.get("slo", "")
        sli_values = svc.get("monthly_sli") or []

        ws.cell(row=row_idx, column=1, value=name)
        ws.cell(row=row_idx, column=2, value=slo)

        # Fill month columns
        for i, label in enumerate(month_labels):
            val = sli_values[i] if i < len(sli_values) else ""
            cell = ws.cell(row=row_idx, column=3 + i, value=val)
            _style_body_cell(cell)

        row_idx += 1

    _auto_width(ws)
    ws.freeze_panes = "A2"


def _build_summary_sheet(wb: Workbook, sla_data: Dict[str, Any]) -> None:
    """
    Optional summary sheet listing SLAs and number of services.
    """
    ws = wb.active
    ws.title = "Summary"

    headers = ["SLA ID", "SLA name", "Services", "Periods"]
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        _style_header_cell(cell)

    row_idx = 2
    for sla_id, info in sla_data.items():
        name = info.get("sla_name") or f"SLA {sla_id}"
        services = info.get("service_data") or {}
        periods = info.get("periods")

        ws.cell(row=row_idx, column=1, value=sla_id)
        ws.cell(row=row_idx, column=2, value=name)
        ws.cell(row=row_idx, column=3, value=len(services))
        ws.cell(row=row_idx, column=4, value=periods)

        for col_idx in range(1, 5):
            _style_body_cell(ws.cell(row=row_idx, column=col_idx))

        row_idx += 1

    _auto_width(ws)
    ws.freeze_panes = "A2"


def create_sla_xlsx_bytes(
    sla_data: Optional[Dict[str, Any]] = None,
    periods: int = 12,
) -> BytesIO:
    """
    Create an XLSX file in memory for SLA/SLI data.

    If sla_data is None, it will call get_sla_sli(periods=periods) itself.

    Returns:
        BytesIO positioned at the beginning, ready to send or save.
    """
    if sla_data is None:
        sla_data = get_sla_sli(periods=periods)

    wb = Workbook()

    # First sheet: summary
    _build_summary_sheet(wb, sla_data)

    # One sheet per SLA
    for sla_id, info in sla_data.items():
        _build_sla_sheet(wb, sla_id, info)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
