# backend/reports/monthly_report.py

from __future__ import annotations

import shutil
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import matplotlib
matplotlib.use("Agg")  # non-interactive backend
import matplotlib.pyplot as plt

from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)

from openpyxl import Workbook

from .. import config
from . import sla as sla_mod
from . import availability as availability_mod
from . import icmp as icmp_mod
from . import firewall_if_usage as fw_usage_mod


# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

BASE_DIR: Path = (config.CACHE_DIR / "customer_report").resolve()
REPORT_DIR: Path = BASE_DIR / "monthly_report"
ZIP_PATH: Path = (config.TMP_DIR / "monthly_report.zip").resolve()

BASE_DIR.mkdir(parents=True, exist_ok=True)
REPORT_DIR.mkdir(parents=True, exist_ok=True)


def _safe_save_fig(path: Path) -> None:
    """
    Save current matplotlib figure to 'path', creating parent directories.
    """
    path.parent.mkdir(parents=True, exist_ok=True)
    plt.tight_layout()
    plt.savefig(path)
    plt.close()


def _build_zip_from_report_dir() -> Path:
    """
    Zip everything under REPORT_DIR into ZIP_PATH.
    """
    if not REPORT_DIR.exists():
        raise RuntimeError(f"Monthly report dir does not exist: {REPORT_DIR}")

    ZIP_PATH.parent.mkdir(parents=True, exist_ok=True)
    tmp_zip = ZIP_PATH.with_suffix(".zip.tmp")

    with zipfile.ZipFile(tmp_zip, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in REPORT_DIR.rglob("*"):
            if path.is_file():
                arcname = path.relative_to(REPORT_DIR)
                zf.write(path, arcname.as_posix())

    tmp_zip.replace(ZIP_PATH)
    return ZIP_PATH


# ---------------------------------------------------------------------------
# Data fetch helpers (use internal modules)
# ---------------------------------------------------------------------------

def _fetch_sla(periods: int = 12) -> Dict[str, Any]:
    """
    SLA/SLI data from backend.reports.sla
    """
    return sla_mod.get_sla_sli(periods=periods)


def _fetch_availability(days: int = 30) -> Dict[str, List[Dict[str, Any]]]:
    """
    Grouped server availability from backend.reports.availability
    Returns:
      { group_name: [ {host, availability, ...}, ... ], ... }
    """
    rows = availability_mod.get_availability(days=days)
    by_group: Dict[str, List[Dict[str, Any]]] = {}

    for row in rows:
        host = row.get("host")
        availability = row.get("availability", "No data")
        groups = row.get("group_names") or ["Ungrouped"]
        for g in groups:
            by_group.setdefault(g, []).append(
                {
                    "group": g,
                    "host": host,
                    "availability": availability,
                    "observed_samples": row.get("observed_samples"),
                    "missing_samples": row.get("missing_samples"),
                    "expected_samples": row.get("expected_samples"),
                }
            )
    return by_group


def _fetch_switch_icmp(days: int = 30) -> Dict[str, Dict[str, Dict[str, Any]]]:
    """
    ICMP history for switches, grouped by host and kind: status/loss/resp.
    """
    docs = icmp_mod.get_icmp_history(days=days)
    mapping: Dict[str, Dict[str, Dict[str, Any]]] = {}

    for doc in docs:
        host = doc.get("host", "")
        if not host.startswith("tvv-sw-"):
            continue

        suffix = doc.get("file_suffix", "")
        if suffix == "ICMP":
            kind = "status"
        elif suffix == "ICMP-loss":
            kind = "loss"
        elif suffix == "ICMP-response-time":
            kind = "resp"
        else:
            continue

        mapping.setdefault(host, {})[kind] = doc

    return mapping


def _fetch_wan_usage(days: int = 30) -> Dict[str, Tuple[List[datetime], List[float]]]:
    """
    WAN usage series per firewall from backend.reports.firewall_if_usage.
    Values converted from bps to Mbps.
    """
    docs = fw_usage_mod.get_firewall_interface_usage(days=days)
    result: Dict[str, Tuple[List[datetime], List[float]]] = {}

    for doc in docs:
        item = doc.get("item", {})
        units = item.get("units")
        if units != "bps":
            continue

        host = doc.get("host") or "unknown"
        hist = doc.get("history", [])

        xs: List[datetime] = []
        ys: List[float] = []

        for h in hist:
            try:
                clk = int(h.get("clock"))
                val = float(h.get("value"))
            except Exception:
                continue
            xs.append(datetime.fromtimestamp(clk, tz=timezone.utc))
            ys.append(val / 1_000_000.0)  # Mbps

        if xs:
            result[host] = (xs, ys)

    return result


# ---------------------------------------------------------------------------
# SLA report (PDF + XLSX)
# ---------------------------------------------------------------------------

def _create_sla_pdf(data: Dict[str, Any], out_dir: Path) -> None:
    pdf = out_dir / "sla_report.pdf"
    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    title = styles["Title"]

    doc = SimpleDocTemplate(str(pdf), pagesize=A4)
    story: List[Any] = []

    story.append(Paragraph("SLA / SLI – 12 months", title))
    story.append(Spacer(1, 18))

    for sla_id, sla_info in data.items():
        sla_name = sla_info.get("sla_name", sla_id)
        story.append(Paragraph(f"<b>SLA:</b> {sla_name}", normal))
        story.append(Spacer(1, 6))

        for sid, svc in sla_info.get("service_data", {}).items():
            sname = svc.get("name", sid)
            slo = svc.get("slo", "N/A")
            months = svc.get("month_labels", [])
            sli = svc.get("monthly_sli", [])

            story.append(Paragraph(f"Service: {sname} (SLO: {slo})", normal))
            line = ", ".join(f"{m}: {v}" for m, v in zip(months, sli))
            story.append(Paragraph(line or "No data", normal))
            story.append(Spacer(1, 4))

        story.append(Spacer(1, 8))

    doc.build(story)


def _create_sla_xlsx(data: Dict[str, Any], out_dir: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "SLA"

    ws.append(["SLA", "Service", "SLO", "Month", "SLI"])

    for sla_id, sla_info in data.items():
        sla_name = sla_info.get("sla_name", sla_id)
        for sid, svc in sla_info.get("service_data", {}).items():
            sname = svc.get("name", sid)
            slo = svc.get("slo", "N/A")
            months = svc.get("month_labels", [])
            sli = svc.get("monthly_sli", [])
            for m, v in zip(months, sli):
                ws.append([sla_name, sname, slo, m, v])

    wb.save(str(out_dir / "sla_report.xlsx"))


# ---------------------------------------------------------------------------
# Server availability / uptime (PDF + XLSX + charts)
# ---------------------------------------------------------------------------

def _summarize_availability(servers: List[Dict[str, Any]]) -> Dict[str, int]:
    summary: Dict[str, int] = {}
    for s in servers:
        avail = str(s.get("availability", "No data"))
        summary[avail] = summary.get(avail, 0) + 1
    return summary


def _create_availability_pdf(by_group: Dict[str, List[Dict[str, Any]]], out_dir: Path) -> None:
    pdf = out_dir / "server_uptime_report.pdf"
    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    title = styles["Title"]

    doc = SimpleDocTemplate(str(pdf), pagesize=A4)
    story: List[Any] = []

    story.append(Paragraph("Server availability – last 30 days", title))
    story.append(Spacer(1, 18))

    for group, servers in by_group.items():
        story.append(Paragraph(f"<b>Group:</b> {group}", normal))
        story.append(Spacer(1, 6))

        summary = _summarize_availability(servers)
        data_tbl = [["Availability", "Hosts"]]
        for avail, count in sorted(summary.items()):
            data_tbl.append([avail, str(count)])

        tbl = Table(data_tbl, hAlign="LEFT", colWidths=[120, 60])
        tbl.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ]
            )
        )
        story.append(tbl)
        story.append(Spacer(1, 12))

    doc.build(story)


def _create_availability_xlsx(by_group: Dict[str, List[Dict[str, Any]]], out_dir: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Server Uptime"

    ws.append(
        [
            "Group",
            "Host",
            "Availability",
            "Observed samples",
            "Missing samples",
            "Expected samples",
        ]
    )

    for group, servers in by_group.items():
        for s in servers:
            ws.append(
                [
                    group,
                    s.get("host"),
                    s.get("availability"),
                    s.get("observed_samples"),
                    s.get("missing_samples"),
                    s.get("expected_samples"),
                ]
            )

    wb.save(str(out_dir / "server_uptime_report.xlsx"))


def _create_availability_charts(by_group: Dict[str, List[Dict[str, Any]]], out_dir: Path) -> None:
    charts_dir = out_dir / "uptime_charts"
    charts_dir.mkdir(parents=True, exist_ok=True)

    for group, servers in by_group.items():
        summary = _summarize_availability(servers)
        labels = sorted(summary.keys())
        counts = [summary[k] for k in labels]

        plt.figure(figsize=(6, 4))
        plt.bar(labels, counts)
        plt.title(f"Availability distribution – {group}")
        plt.xlabel("Availability")
        plt.ylabel("Hosts")
        for i, cnt in enumerate(counts):
            plt.text(i, cnt + 0.1, str(cnt), ha="center", va="bottom", fontsize=8)

        out_path = charts_dir / f"{group}_availability.png"
        _safe_save_fig(out_path)


# ---------------------------------------------------------------------------
# Switch ICMP (XLSX + graphs)
# ---------------------------------------------------------------------------

def _create_switch_icmp_xlsx(
    switches: Dict[str, Dict[str, Dict[str, Any]]], out_dir: Path
) -> None:
    if not switches:
        return

    out_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Switch ICMP"

    ws.append(["Switch", "Last status value", "Last loss (%)", "Last response time (sec)"])

    def last_val(doc: Optional[Dict[str, Any]]) -> Optional[float]:
        if not doc:
            return None
        hist = doc.get("history") or []
        if not hist:
            return None
        try:
            return float(hist[-1].get("value"))
        except Exception:
            return None

    for sw, kinds in sorted(switches.items()):
        ws.append(
            [
                sw,
                last_val(kinds.get("status")) or "",
                last_val(kinds.get("loss")) or "",
                last_val(kinds.get("resp")) or "",
            ]
        )

    wb.save(str(out_dir / "switch_icmp_report.xlsx"))


def _create_switch_icmp_graphs(
    switches: Dict[str, Dict[str, Dict[str, Any]]], out_dir: Path
) -> None:
    if not switches:
        return

    out_dir.mkdir(parents=True, exist_ok=True)

    def series_from_doc(doc: Dict[str, Any]) -> Tuple[List[datetime], List[float]]:
        xs: List[datetime] = []
        ys: List[float] = []
        for h in doc.get("history", []):
            try:
                clk = int(h.get("clock"))
                val = float(h.get("value"))
            except Exception:
                continue
            xs.append(datetime.fromtimestamp(clk, tz=timezone.utc))
            ys.append(val)
        return xs, ys

    # Per-switch graphs
    for sw, kinds in sorted(switches.items()):
        loss_doc = kinds.get("loss")
        if loss_doc:
            xs, ys = series_from_doc(loss_doc)
            if xs:
                plt.figure(figsize=(8, 3))
                plt.plot(xs, ys)
                plt.title(f"{sw} ICMP loss (%)")
                plt.xlabel("Time (UTC)")
                plt.ylabel("Loss %")
                _safe_save_fig(out_dir / f"{sw}_icmp_loss.png")

        resp_doc = kinds.get("resp")
        if resp_doc:
            xs, ys = series_from_doc(resp_doc)
            if xs:
                plt.figure(figsize=(8, 3))
                plt.plot(xs, ys)
                plt.title(f"{sw} ICMP response time (sec)")
                plt.xlabel("Time (UTC)")
                plt.ylabel("Seconds")
                _safe_save_fig(out_dir / f"{sw}_icmp_response_time.png")

    # Aggregated loss
    plt.figure(figsize=(10, 4))
    any_line = False
    for sw, kinds in sorted(switches.items()):
        loss_doc = kinds.get("loss")
        if not loss_doc:
            continue
        xs, ys = series_from_doc(loss_doc)
        if xs:
            plt.plot(xs, ys, label=sw)
            any_line = True
    if any_line:
        plt.title("ICMP loss (%) – all switches")
        plt.xlabel("Time (UTC)")
        plt.ylabel("Loss %")
        plt.legend(ncol=2, fontsize="small")
        _safe_save_fig(out_dir / "all-switches_icmp_loss.png")
    else:
        plt.close()

    # Aggregated response time
    plt.figure(figsize=(10, 4))
    any_line = False
    for sw, kinds in sorted(switches.items()):
        resp_doc = kinds.get("resp")
        if not resp_doc:
            continue
        xs, ys = series_from_doc(resp_doc)
        if xs:
            plt.plot(xs, ys, label=sw)
            any_line = True
    if any_line:
        plt.title("ICMP response time (sec) – all switches")
        plt.xlabel("Time (UTC)")
        plt.ylabel("Seconds")
        plt.legend(ncol=2, fontsize="small")
        _safe_save_fig(out_dir / "all-switches_icmp_response_time.png")
    else:
        plt.close()


# ---------------------------------------------------------------------------
# WAN usage (XLSX + graphs)
# ---------------------------------------------------------------------------

def _create_wan_usage_xlsx(
    wan: Dict[str, Tuple[List[datetime], List[float]]], out_dir: Path
) -> None:
    if not wan:
        return

    out_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "WAN Usage"

    ws.append(["Firewall", "Timestamp UTC", "Mbps"])

    for fw, (xs, ys) in sorted(wan.items()):
        for dt, mbps in zip(xs, ys):
            ws.append([fw, dt.isoformat(), mbps])

    wb.save(str(out_dir / "wan_usage_report.xlsx"))


def _create_wan_usage_graphs(
    wan: Dict[str, Tuple[List[datetime], List[float]]], out_dir: Path
) -> None:
    if not wan:
        return

    out_dir.mkdir(parents=True, exist_ok=True)

    # Per-firewall graphs
    for fw, (xs, ys) in sorted(wan.items()):
        if not xs:
            continue
        plt.figure(figsize=(8, 3))
        plt.plot(xs, ys)
        plt.title(f"Usage WAN Service – {fw} (Mbps)")
        plt.xlabel("Time (UTC)")
        plt.ylabel("Mbps")
        _safe_save_fig(out_dir / f"{fw}_wan_usage_mbps.png")

    # Aggregated graph
    plt.figure(figsize=(10, 4))
    any_line = False
    for fw, (xs, ys) in sorted(wan.items()):
        if xs:
            plt.plot(xs, ys, label=fw)
            any_line = True
    if any_line:
        plt.title("Usage WAN Service – all firewalls (Mbps)")
        plt.xlabel("Time (UTC)")
        plt.ylabel("Mbps")
        plt.legend(ncol=2, fontsize="small")
        _safe_save_fig(out_dir / "all-firewalls_wan_usage_mbps.png")
    else:
        plt.close()


# ---------------------------------------------------------------------------
# Public API for FastAPI
# ---------------------------------------------------------------------------

def generate_monthly_report() -> Path:
    """
    Main entrypoint:
      1) Clear previous monthly_report directory.
      2) Generate SLA reports (PDF + XLSX).
      3) Generate server availability reports (PDF + XLSX + charts).
      4) Generate switch ICMP XLSX + graphs.
      5) Generate WAN usage XLSX + graphs.
      6) Build and return monthly_report.zip.
    """
    if REPORT_DIR.exists():
        shutil.rmtree(REPORT_DIR)
    REPORT_DIR.mkdir(parents=True, exist_ok=True)

    # 1) SLA
    sla_data = _fetch_sla(periods=12)
    _create_sla_pdf(sla_data, REPORT_DIR)
    _create_sla_xlsx(sla_data, REPORT_DIR)

    # 2) Server availability
    avail = _fetch_availability(days=30)
    _create_availability_pdf(avail, REPORT_DIR)
    _create_availability_xlsx(avail, REPORT_DIR)
    _create_availability_charts(avail, REPORT_DIR)

    # 3) Switch ICMP
    switches = _fetch_switch_icmp(days=30)
    if switches:
        sw_dir = REPORT_DIR / "switch-icmp"
        _create_switch_icmp_xlsx(switches, sw_dir)
        _create_switch_icmp_graphs(switches, sw_dir)

    # 4) WAN usage
    wan = _fetch_wan_usage(days=30)
    if wan:
        wan_dir = REPORT_DIR / "wan-usage"
        _create_wan_usage_xlsx(wan, wan_dir)
        _create_wan_usage_graphs(wan, wan_dir)

    return _build_zip_from_report_dir()


def get_existing_report_path() -> Optional[Path]:
    if ZIP_PATH.exists():
        return ZIP_PATH
    return None
