import os
import csv
from collections import defaultdict
from datetime import datetime

import openpyxl
import networkx as nx
import xml.etree.ElementTree as ET
import uuid

from report_config import OUTPUT_DIR, EXCLUDED_HOSTS


def write_summary_excel(rows, suffix: str = "") -> None:
    fname = f"network_blueprint_summary{suffix}.xlsx"
    path = os.path.join(OUTPUT_DIR, fname)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(
        [
            "ZabbixHost",
            "ConnectionType",
            "LocalIP",
            "LocalPort",
            "RemoteIP",
            "RemotePort",
            "Count",
            "RemoteHostName",
            "LatestTimestamp",
        ]
    )
    for r in rows:
        ws.append(
            [
                r["local_host"],
                r["type"],
                r["local_ip"],
                r.get("port", ""),
                r["remote_ip"],
                r.get("port", ""),
                r.get("count", 0),
                r["remote_host"],
                r["timestamp"],
            ]
        )
    wb.save(path)
    print(f"[REPORT] Summary Excel: {path}")


def write_per_host_excel(rows, suffix: str = "") -> None:
    fname = f"network_blueprint_per_host{suffix}.xlsx"
    path = os.path.join(OUTPUT_DIR, fname)
    wb = openpyxl.Workbook()
    per = defaultdict(list)
    for r in rows:
        per[r["local_host"]].append(r)
    wb.remove(wb.active)
    for host, recs in per.items():
        ws = wb.create_sheet(title=host[:31])
        ws.append(
            [
                "ZabbixHost",
                "Type",
                "LocalIP",
                "Port",
                "RemoteIP",
                "RemoteHost",
                "Count",
                "LatestTimestamp",
            ]
        )
        for r in recs:
            ws.append(
                [
                    r["local_host"],
                    r["type"],
                    r["local_ip"],
                    r.get("port", ""),
                    r["remote_ip"],
                    r["remote_host"],
                    r.get("count", 0),
                    r["timestamp"],
                ]
            )
    wb.save(path)
    print(f"[REPORT] Per-host Excel: {path}")


def write_gephi_csv(rows, suffix: str = "") -> None:
    fname = f"network_blueprint_gephi{suffix}.csv"
    path = os.path.join(OUTPUT_DIR, fname)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Source", "SourceIP", "Target", "TargetIP", "Port", "Count"])
        for r in rows:
            if not r["local_host"] or not r["remote_host"]:
                continue
            if r["type"] == "outgoing":
                src, sip = r["local_host"], r["local_ip"]
                dst, dip = r["remote_host"], r["remote_ip"]
            else:
                src, sip = r["remote_host"], r["remote_ip"]
                dst, dip = r["local_host"], r["local_ip"]
            w.writerow([src, sip, dst, dip, r.get("port", ""), r.get("count", 0)])
    print(f"[REPORT] CSV: {path}")


def separate_overlaps(
    node_positions,
    node_width: int = 160,
    node_height: int = 80,
    padding: int = 40,
    max_iterations: int = 800,
) -> None:
    nodes = list(node_positions.keys())

    def overlap(a, b):
        return not (a[1] < b[0] or a[0] > b[1] or a[3] < b[2] or a[2] > b[3])

    for _ in range(max_iterations):
        moved = False
        for i in range(len(nodes)):
            for j in range(i + 1, len(nodes)):
                n1, n2 = nodes[i], nodes[j]
                x1, y1 = node_positions[n1]
                x2, y2 = node_positions[n2]
                box1 = [x1, x1 + node_width, y1, y1 + node_height]
                box2 = [x2, x2 + node_width, y2, y2 + node_height]
                if overlap(box1, box2):
                    moved = True
                    dx, dy = x2 - x1, y2 - y1
                    if dx == 0 and dy == 0:
                        dx = 1
                    ox = (node_width + padding) - abs(dx)
                    oy = (node_height + padding) - abs(dy)
                    px = (
                        ox / 2 * (1 if dx > 0 else -1)
                        if abs(dx) < (node_width + padding)
                        else 0
                    )
                    py = (
                        oy / 2 * (1 if dy > 0 else -1)
                        if abs(dy) < (node_height + padding)
                        else 0
                    )
                    node_positions[n1] = (x1 - px, y1 - py)
                    node_positions[n2] = (x2 + px, y2 + py)
        if not moved:
            break


def build_host_ip_map(rows):
    host_ips = defaultdict(set)
    for r in rows:
        lh, lip = r["local_host"], r["local_ip"]
        rh, rip = r["remote_host"], r["remote_ip"]
        if lh and lip and lip != lh:
            host_ips[lh].add(lip)
        if rh and rip and rip != rh:
            host_ips[rh].add(rip)
    return {k: sorted(v) for k, v in host_ips.items()}


def build_drawio_per_host(rows, suffix: str = "") -> None:
    fname = f"network_blueprint_per_host{suffix}.drawio"
    path = os.path.join(OUTPUT_DIR, fname)
    per_rows = defaultdict(list)
    for r in rows:
        per_rows[r["local_host"]].append(r)
    host_ips = build_host_ip_map(rows)

    mxfile = ET.Element(
        "mxfile",
        {
            "host": "app.diagrams.net",
            "modified": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
            "agent": "python",
            "version": "15.8.7",
            "type": "device",
        },
    )
    for host, recs in per_rows.items():
        if host in EXCLUDED_HOSTS:
            continue
        G = nx.DiGraph()
        for r in recs:
            if r["remote_host"] in EXCLUDED_HOSTS:
                continue
            src, dst = (
                (r["local_host"], r["remote_host"])
                if r["type"] == "outgoing"
                else (r["remote_host"], r["local_host"])
            )
            G.add_node(src)
            G.add_node(dst)
            lbl = (
                f"{r['type']} (port={r.get('port','')})"
                if r.get("port")
                else r["type"]
            )
            G.add_edge(src, dst, label=lbl)
        if G.number_of_nodes() == 0:
            continue

        pos = nx.spring_layout(G, k=3.0, iterations=200, seed=42)
        scale = 800.0
        xs, ys = zip(*pos.values())
        minx, miny = min(xs), min(ys)
        node_positions = {
            n: ((x - minx) * scale + 50, (y - miny) * scale + 50)
            for n, (x, y) in pos.items()
        }
        separate_overlaps(node_positions)

        diagram = ET.SubElement(
            mxfile,
            "diagram",
            {"id": str(uuid.uuid4())[:8], "name": host[:31]},
        )
        model = ET.SubElement(
            diagram,
            "mxGraphModel",
            {
                "dx": "1600",
                "dy": "1200",
                "grid": "1",
                "gridSize": "10",
                "guides": "1",
                "tooltips": "1",
                "connect": "1",
                "arrows": "1",
                "fold": "1",
                "page": "1",
                "pageScale": "1",
                "pageWidth": "4000",
                "pageHeight": "4000",
                "math": "0",
                "shadow": "0",
            },
        )
        root = ET.SubElement(model, "root")
        ET.SubElement(root, "mxCell", {"id": "0"})
        ET.SubElement(root, "mxCell", {"id": "1", "parent": "0"})

        id_map = {}
        idx = 1
        for n in G.nodes():
            node_id = f"node{idx}"
            id_map[n] = node_id
            ips = host_ips.get(n, [])
            label = f"{n} ({', '.join(ips)})" if ips else n
            cell = ET.SubElement(
                root,
                "mxCell",
                {
                    "id": node_id,
                    "value": label,
                    "style": "shape=rectangle;whiteSpace=wrap;html=1;strokeWidth=2;align=center;",
                    "vertex": "1",
                    "parent": "1",
                },
            )
            x, y = node_positions[n]
            ET.SubElement(
                cell,
                "mxGeometry",
                {
                    "x": str(x),
                    "y": str(y),
                    "width": "160",
                    "height": "80",
                    "as": "geometry",
                },
            )
            idx += 1

        eid = 1
        for src, dst, data in G.edges(data=True):
            edge_id = f"edge{eid}"
            edge = ET.SubElement(
                root,
                "mxCell",
                {
                    "id": edge_id,
                    "value": data.get("label", ""),
                    "style": "endArrow=classic;html=1;",
                    "edge": "1",
                    "parent": "1",
                    "source": id_map[src],
                    "target": id_map[dst],
                },
            )
            ET.SubElement(edge, "mxGeometry", {"relative": "1", "as": "geometry"})
            eid += 1

    tree = ET.ElementTree(mxfile)
    if hasattr(ET, "indent"):
        ET.indent(tree, space="  ", level=0)
    tree.write(path, encoding="utf-8", xml_declaration=True)
    print(f"[REPORT] DrawIO: {path}")
