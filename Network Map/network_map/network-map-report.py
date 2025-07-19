#!/usr/bin/env python3
import os
import json
import requests
import openpyxl
import warnings
import urllib3
import csv
from collections import defaultdict
import networkx as nx
import xml.etree.ElementTree as ET
import uuid
import time
from datetime import datetime
import ipaddress

# Suppress SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
warnings.simplefilter("ignore", urllib3.exceptions.SubjectAltNameWarning)

# ---------------------------
# Output directory
# ---------------------------
OUTPUT_DIR = "/opt/network_map/reports"
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ---------------------------
# Zabbix API Configuration
# ---------------------------
ZABBIX_URL = "https://zabbix.example.se/api_jsonrpc.php"
API_TOKEN = "this_is_a_fake_token_for_example_purposes"

# Report time window: last 30 days
DAYS = 30
TIME_TILL = int(time.time())
TIME_FROM = TIME_TILL - DAYS * 24 * 3600
# Chunk size for history.get (1 day)
HISTORY_CHUNK = 1 * 24 * 3600

# Optionally exclude certain hosts
EXCLUDED_HOSTS = {
    "Zabbix server",
}

# ---------------------------
# NetBox Configuration
# ---------------------------
NETBOX_URL = "https://netbox.example.se"
NETBOX_TOKEN = "this_is_a_fake_token_for_example_purposes"
NETBOX_HEADERS = {
    "Authorization": f"Token {NETBOX_TOKEN}",
    "Content-Type": "application/json",
    "Accept": "application/json",
}

# Internal IP ranges
INTERNAL_NETWORKS = [
    ipaddress.ip_network("10.0.0.0/8"),
    ipaddress.ip_network("192.168.0.0/16"),
    ipaddress.ip_network("172.16.0.0/12"),
]

def zabbix_api(method, params=None):
    headers = {"Content-Type": "application/json-rpc"}
    payload = {
        "jsonrpc": "2.0",
        "method": method,
        "params": params or {},
        "auth": API_TOKEN,
        "id": 1
    }
    resp = requests.post(ZABBIX_URL, headers=headers, json=payload, verify=False)
    if resp.status_code != 200:
        print(f"[ERROR] HTTP {resp.status_code} from Zabbix API: {resp.text}")
        resp.raise_for_status()
    data = resp.json()
    if "error" in data:
        raise Exception(f"Zabbix API error: {data['error']}")
    return data["result"]

def get_netbox_name_for_ip(ip):
    url = f"{NETBOX_URL}/api/ipam/ip-addresses/?address={ip}/32"
    resp = requests.get(url, headers=NETBOX_HEADERS, verify=False)
    resp.raise_for_status()
    results = resp.json().get("results", [])
    if not results:
        return None
    obj = results[0]
    vm = obj.get("assigned_object", {}).get("virtual_machine", {}).get("name")
    if vm:
        return vm
    return obj.get("dns_name")

def get_all_hosts_ip_map():
    params = {
        "output": ["hostid", "host"],
        "selectInterfaces": ["ip"],
        "filter": {"status": 0}
    }
    hosts = zabbix_api("host.get", params)
    ip_to_host = {}
    for h in hosts:
        name = h["host"]
        for iface in h.get("interfaces", []):
            ip = iface.get("ip")
            if ip:
                ip_to_host[ip] = name
    # resolve DNS names via NetBox for any IPs still named by IP
    for ip, name in list(ip_to_host.items()):
        if name == ip:
            nb = get_netbox_name_for_ip(ip)
            if nb:
                ip_to_host[ip] = nb
    return ip_to_host

def get_network_connection_items():
    params = {
        "output": ["itemid"],
        "filter": {"name": [
            "linux-network-connections",
            "windows-network-connections"
        ]},
        "selectHosts": ["host"]
    }
    return zabbix_api("item.get", params)

def get_connection_history(itemid):
    all_entries = []
    for start in range(TIME_FROM, TIME_TILL, HISTORY_CHUNK):
        end = min(start + HISTORY_CHUNK, TIME_TILL)
        params = {
            "output": "extend",
            "history": 4,
            "itemids": itemid,
            "time_from": start,
            "time_till": end,
            "sortfield": "clock",
            "sortorder": "ASC",
            "limit": 100000
        }
        try:
            chunk = zabbix_api("history.get", params)
        except Exception as e:
            print(f"[WARN] Failed history.get for item {itemid} from {start} to {end}: {e}")
            continue
        all_entries.extend(chunk)
    return all_entries

def parse_history_connections(items, ip_to_host):
    agg = {}
    for itm in items:
        host = itm.get("hosts", [{}])[0].get("host", "")
        itemid = itm.get("itemid")
        history = get_connection_history(itemid)
        for entry in history:
            raw = entry.get("value") or ""
            try:
                data = json.loads(raw)
            except json.JSONDecodeError:
                continue
            timestamp = int(entry.get("clock", 0))
            inc = data.get("incomingconnections", []) or []
            out = data.get("outgoingconnections", []) or []
            if isinstance(inc, dict): inc = [inc]
            if isinstance(out, dict): out = [out]
            for conn_list, ctype in ((inc, "incoming"), (out, "outgoing")):
                for c in conn_list:
                    rip = c.get("remoteip", "")
                    local_ip = c.get("localip", "")
                    port = c.get("localport" if ctype=="incoming" else "remoteport", "")
                    remote_host = ip_to_host.get(rip, rip)
                    key = (ctype, host, local_ip, remote_host, rip, port)
                    if key not in agg:
                        agg[key] = {"count": 0, "latest_ts": 0}
                    agg[key]["count"] += 1
                    if timestamp > agg[key]["latest_ts"]:
                        agg[key]["latest_ts"] = timestamp
    rows = []
    for (ctype, local_host, local_ip, remote_host, remote_ip, port), v in agg.items():
        rows.append({
            "type": ctype,
            "local_host": local_host,
            "local_ip": local_ip,
            "remote_host": remote_host,
            "remote_ip": remote_ip,
            "port": port,
            "count": v["count"],
            "timestamp": datetime.utcfromtimestamp(v["latest_ts"]).isoformat() + 'Z'
        })
    return rows

def filter_internal(rows):
    filtered = []
    for r in rows:
        try:
            lip = ipaddress.ip_address(r["local_ip"])
            rip = ipaddress.ip_address(r["remote_ip"])
        except ValueError:
            continue
        if any(lip in net for net in INTERNAL_NETWORKS) and any(rip in net for net in INTERNAL_NETWORKS):
            filtered.append(r)
    return filtered

def filter_public(rows):
    filtered = []
    for r in rows:
        try:
            lip = ipaddress.ip_address(r["local_ip"])
            rip = ipaddress.ip_address(r["remote_ip"])
        except ValueError:
            continue
        if lip.is_loopback or rip.is_loopback:
            continue
        if not any(rip in net for net in INTERNAL_NETWORKS):
            filtered.append(r)
    return filtered

def write_summary_excel(rows, suffix=""):
    fname = f"network_blueprint_summary{suffix}.xlsx"
    path = os.path.join(OUTPUT_DIR, fname)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ZabbixHost", "ConnectionType", "LocalIP", "LocalPort",
               "RemoteIP", "RemotePort", "Count", "RemoteHostName", "LatestTimestamp"])
    for r in rows:
        ws.append([
            r["local_host"], r["type"], r["local_ip"], r.get("port", ""),
            r["remote_ip"], r.get("port", ""), r.get("count", 0), r["remote_host"], r["timestamp"]
        ])
    wb.save(path)
    print(f"Summary Excel: {path}")

def write_per_host_excel(rows, suffix=""):
    fname = f"network_blueprint_per_host{suffix}.xlsx"
    path = os.path.join(OUTPUT_DIR, fname)
    wb = openpyxl.Workbook()
    per = defaultdict(list)
    for r in rows:
        per[r["local_host"]].append(r)
    wb.remove(wb.active)
    for host, recs in per.items():
        ws = wb.create_sheet(title=host[:31])
        ws.append(["ZabbixHost", "Type", "LocalIP", "Port", "RemoteIP", "RemoteHost", "Count", "LatestTimestamp"])
        for r in recs:
            ws.append([
                r["local_host"], r["type"], r["local_ip"], r.get("port", ""),
                r["remote_ip"], r["remote_host"], r.get("count", 0), r["timestamp"]
            ])
    wb.save(path)
    print(f"Per-host Excel: {path}")

def write_gephi_csv(rows, suffix=""):
    fname = f"network_blueprint_gephi{suffix}.csv"
    path = os.path.join(OUTPUT_DIR, fname)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Source", "SourceIP", "Target", "TargetIP", "Port", "Count"])
        for r in rows:
            if not r["local_host"] or not r["remote_host"]:
                continue
            if r["type"] == "outgoing":
                src, sip = r["local_host"], r["local_ip"]
                dst, dip = r["remote_host"], r["remote_ip"]
            else:
                src, sip = r["remote_host"], r["remote_ip"]
                dst, dip = r["local_host"], r["local_ip"]
            w.writerow([src, sip, dst, dip, r.get("port", ""), r.get("count", 0)])
    print(f"CSV file created: {path}")

def separate_overlaps(node_positions, node_width=160, node_height=80, padding=40, max_iterations=800):
    nodes = list(node_positions.keys())
    def overlap(a, b):
        return not (a[1] < b[0] or a[0] > b[1] or a[3] < b[2] or a[2] > b[3])
    for _ in range(max_iterations):
        moved = False
        for i in range(len(nodes)):
            for j in range(i+1, len(nodes)):
                n1, n2 = nodes[i], nodes[j]
                x1, y1 = node_positions[n1]; x2, y2 = node_positions[n2]
                box1 = [x1, x1+node_width, y1, y1+node_height]
                box2 = [x2, x2+node_width, y2, y2+node_height]
                if overlap(box1, box2):
                    moved = True
                    dx, dy = x2-x1, y2-y1
                    if dx==0 and dy==0: dx=1
                    ox = (node_width+padding)-abs(dx)
                    oy = (node_height+padding)-abs(dy)
                    px = ox/2*(1 if dx>0 else -1) if abs(dx)<(node_width+padding) else 0
                    py = oy/2*(1 if dy>0 else -1) if abs(dy)<(node_height+padding) else 0
                    node_positions[n1] = (x1-px, y1-py)
                    node_positions[n2] = (x2+px, y2+py)
        if not moved:
            break

def build_host_ip_map(rows):
    host_ips = defaultdict(set)
    for r in rows:
        lh, lip = r["local_host"], r["local_ip"]
        rh, rip = r["remote_host"], r["remote_ip"]
        if lh and lip and lip != lh:
            host_ips[lh].add(lip)
        if rh and rip and rip != rh:
            host_ips[rh].add(rip)
    return {k: sorted(v) for k, v in host_ips.items()}

def build_drawio_per_host(rows, suffix=""):
    fname = f"network_blueprint_per_host{suffix}.drawio"
    path = os.path.join(OUTPUT_DIR, fname)
    per_rows = defaultdict(list)
    for r in rows:
        per_rows[r["local_host"]].append(r)
    host_ips = build_host_ip_map(rows)

    mxfile = ET.Element(
        "mxfile",
        {
            "host": "app.diagrams.net",
            "modified": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
            "agent": "python",
            "version": "15.8.7",
            "type": "device"
        }
    )
    for host, recs in per_rows.items():
        if host in EXCLUDED_HOSTS:
            continue
        G = nx.DiGraph()
        for r in recs:
            if r["remote_host"] in EXCLUDED_HOSTS:
                continue
            src, dst = (
                (r["local_host"], r["remote_host"])
                if r["type"] == "outgoing"
                else (r["remote_host"], r["local_host"])
            )
            G.add_node(src)
            G.add_node(dst)
            lbl = (
                f"{r['type']} (port={r.get('port','')})"
                if r.get('port')
                else r['type']
            )
            G.add_edge(src, dst, label=lbl)
        if G.number_of_nodes() == 0:
            continue

        pos = nx.spring_layout(G, k=3.0, iterations=200, seed=42)
        scale = 800.0
        xs, ys = zip(*pos.values())
        minx, miny = min(xs), min(ys)
        node_positions = {
            n: ((x - minx) * scale + 50, (y - miny) * scale + 50)
            for n, (x, y) in pos.items()
        }
        separate_overlaps(node_positions)

        diagram = ET.SubElement(
            mxfile,
            "diagram",
            {"id": str(uuid.uuid4())[:8], "name": host[:31]}
        )
        model = ET.SubElement(
            diagram,
            "mxGraphModel",
            {
                "dx": "1600", "dy": "1200", "grid": "1", "gridSize": "10",
                "guides": "1", "tooltips": "1", "connect": "1", "arrows": "1",
                "fold": "1", "page": "1", "pageScale": "1", "pageWidth": "4000",
                "pageHeight": "4000", "math": "0", "shadow": "0"
            }
        )
        root = ET.SubElement(model, "root")
        ET.SubElement(root, "mxCell", {"id": "0"})
        ET.SubElement(root, "mxCell", {"id": "1", "parent": "0"})

        id_map = {}
        idx = 1
        for n in G.nodes():
            node_id = f"node{idx}"
            id_map[n] = node_id
            ips = host_ips.get(n, [])
            label = f"{n} ({', '.join(ips)})" if ips else n
            cell = ET.SubElement(
                root,
                "mxCell",
                {
                    "id": node_id,
                    "value": label,
                    "style": "shape=rectangle;whiteSpace=wrap;html=1;strokeWidth=2;align=center;",
                    "vertex": "1",
                    "parent": "1"
                }
            )
            x, y = node_positions[n]
            ET.SubElement(
                cell,
                "mxGeometry",
                {"x": str(x), "y": str(y), "width": "160", "height": "80", "as": "geometry"}
            )
            idx += 1

        eid = 1
        for src, dst, data in G.edges(data=True):
            edge_id = f"edge{eid}"
            edge = ET.SubElement(
                root,
                "mxCell",
                {
                    "id": edge_id,
                    "value": data.get("label", ""),
                    "style": "endArrow=classic;html=1;",
                    "edge": "1",
                    "parent": "1",
                    "source": id_map[src],
                    "target": id_map[dst]
                }
            )
            ET.SubElement(edge, "mxGeometry", {"relative": "1", "as": "geometry"})
            eid += 1

    tree = ET.ElementTree(mxfile)
    ET.indent(tree, space="  ", level=0)
    tree.write(path, encoding="utf-8", xml_declaration=True)
    print(f"DrawIO file created: {path}")

def main():
    ip_map = get_all_hosts_ip_map()
    items = get_network_connection_items()
    if not items:
        print("No network‑connection items found. Exiting.")
        return

    rows = parse_history_connections(items, ip_map)

    # Full exports
    write_summary_excel(rows, suffix="")
    write_per_host_excel(rows, suffix="")
    write_gephi_csv(rows, suffix="")
    build_drawio_per_host(rows, suffix="")

    # Internal‑only exports
    internal_rows = filter_internal(rows)
    write_summary_excel(internal_rows,  suffix="_internal_ip")
    write_per_host_excel(internal_rows,  suffix="_internal_ip")
    write_gephi_csv(internal_rows,       suffix="_internal_ip")
    build_drawio_per_host(internal_rows, suffix="_internal_ip")

    # Public‑only exports
    public_rows = filter_public(rows)
    write_summary_excel(public_rows,  suffix="_public_ip")
    write_per_host_excel(public_rows,  suffix="_public_ip")
    write_gephi_csv(public_rows,       suffix="_public_ip")
    build_drawio_per_host(public_rows, suffix="_public_ip")

if __name__ == "__main__":
    main()
